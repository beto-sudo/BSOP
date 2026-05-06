#!/usr/bin/env python3
"""
Importa el padrón de proveedores de DILESA exportado de CONTPAQi
(`Padrón de Proveedores.xlsx`) al schema actual: `erp.personas` +
`erp.proveedores`.

Estrategia mínima, una sola pasada:

1. Lee el Excel con `openpyxl` (mismo patrón que `import-empleados-contpaqi`).
2. Filtra a las filas que son proveedores (col A = código numérico).
3. Emite **un solo SQL transaccional** a stdout (o `--out`) con:
   - CTE `excel(codigo, nombre, rfc, tipo_persona)` con los 196 valores.
   - INSERT INTO `erp.personas` para los RFCs que NO existen en DILESA
     (match por RFC normalizado: replace('-','')).
   - INSERT INTO `erp.proveedores` vinculando a la persona DILESA
     (existente o recién creada) para los RFCs que NO son ya proveedor.
   - Resumen final con SELECTs.

**Datos contables ignorados** (cuenta `201-01-001`, tasas IVA/ISR,
concepto IETU). Diferidos a futura iniciativa de contabilidad.

**Dedup cross-empresa**: los RFCs que existen solo en otra empresa
(p.ej. RDB) generan una NUEVA persona DILESA — instrucción de Beto:
"una persona por empresa".

**Uso:**

```bash
EXCEL=~/Downloads/'Padron de Proveedores.xlsx' \\
  python3 scripts/import-contpaqi/dilesa-proveedores-import.py \\
    > scripts/import-contpaqi/sql/apply-dilesa-proveedores-2026-05-06.sql
```

Después aplicar manualmente via MCP `apply_migration` o `psql`.
"""

from __future__ import annotations

import os
import sys
from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook


def parse_tasa(raw: str | None) -> float | None:
    """' 8%' -> 0.08, '16%' -> 0.16, ' 8% 0%' -> 0.08 (primera), None/inválido -> None."""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    first_token = s.split()[0]
    if not first_token.endswith("%"):
        return None
    try:
        pct = float(first_token.rstrip("%"))
    except ValueError:
        return None
    val = round(pct / 100, 4)
    if val in (0.0, 0.08, 0.16):
        return val
    return None


def read_excel(xlsx_path: Path) -> Iterator[tuple[str, str, str, float | None]]:
    """Yields (codigo, nombre, rfc, tasa_iva) for each provider.

    El Excel CONTPAQi pone cada proveedor en 2 filas:
    - Fila A: codigo, nombre, rfc, curp, cuenta, tipo_tercero, tipo_operacion
    - Fila B: (vacío en A,B), id_fiscal, extranjero, pais, nacionalidad, tasas, ...

    Recolectamos ambas y emitimos un solo registro por par.
    """
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    pending: tuple[str, str, str] | None = None
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 7:
            if pending:
                yield (pending[0], pending[1], pending[2], None)
                pending = None
            continue
        codigo, nombre, rfc = row[0], row[1], row[2]
        if codigo is not None:
            if pending:
                yield (pending[0], pending[1], pending[2], None)
                pending = None
            try:
                codigo_int = int(float(codigo))
            except (TypeError, ValueError):
                continue
            if nombre is None or rfc is None:
                continue
            rfc_str = str(rfc).strip()
            if len(rfc_str) not in (12, 13):
                continue
            pending = (str(codigo_int), str(nombre).strip(), rfc_str)
        else:
            if pending:
                tasa = parse_tasa(row[6]) if len(row) > 6 else None
                yield (pending[0], pending[1], pending[2], tasa)
                pending = None
    if pending:
        yield (pending[0], pending[1], pending[2], None)


def sql_quote(s: str) -> str:
    return "'" + s.replace("'", "''") + "'"


def derive_tipo_persona(rfc: str) -> str:
    return "moral" if len(rfc) == 12 else "fisica"


def main() -> int:
    xlsx_default = Path.home() / "Downloads" / "Padron de Proveedores.xlsx"
    xlsx = Path(os.environ.get("EXCEL", str(xlsx_default)))
    if not xlsx.exists():
        print(f"ERROR: archivo no encontrado: {xlsx}", file=sys.stderr)
        return 1

    rows = list(read_excel(xlsx))
    seen: dict[str, tuple[str, str, float | None]] = {}
    for codigo, nombre, rfc, tasa in rows:
        if rfc not in seen:
            seen[rfc] = (codigo, nombre, tasa)
    sin_tasa = sum(1 for _, _, tasa in seen.values() if tasa is None)
    print(
        f"-- Leídas {len(rows)} filas, {len(seen)} RFCs únicos, {sin_tasa} sin tasa válida",
        file=sys.stderr,
    )

    def fmt_tasa(t: float | None) -> str:
        return "NULL" if t is None else f"{t}"

    values = [
        f"({sql_quote(codigo)}, {sql_quote(nombre)}, {sql_quote(rfc)}, "
        f"{sql_quote(derive_tipo_persona(rfc))}, {fmt_tasa(tasa)})"
        for rfc, (codigo, nombre, tasa) in sorted(seen.items())
    ]

    out = sys.stdout

    out.write(
        "-- Generado por scripts/import-contpaqi/dilesa-proveedores-import.py\n"
        "-- Fecha: 2026-05-06\n"
        "-- Fuente: ~/Downloads/Padron de Proveedores.xlsx (CONTPAQi DILESA)\n"
        "--\n"
        "-- Aplica:\n"
        "--   1. Alta de personas DILESA para los RFCs que NO existen en DILESA.\n"
        "--   2. Alta de proveedores para los RFCs que aún no son proveedores.\n"
        "--\n"
        "-- Match por RFC normalizado (sin guiones) — la DB legacy tiene\n"
        "-- algunos RFCs con guión (ej. CACX-870331-S60). El INSERT NO\n"
        "-- normaliza el RFC nuevo (los del Excel ya vienen sin guión).\n"
        "--\n"
        "-- Cross-empresa: RFCs que solo existen en otra empresa generan\n"
        "-- nueva persona DILESA (instrucción Beto: una persona por empresa).\n"
        "--\n"
        "-- Datos contables (cuenta CONTPAQi, tasas IVA/ISR, concepto IETU)\n"
        "-- diferidos. El `codigo` CONTPAQi se persiste en `erp.proveedores.codigo`\n"
        "-- como referencia para futura iniciativa de contabilidad.\n\n"
    )

    out.write("BEGIN;\n\n")

    out.write("WITH excel(codigo, nombre, rfc, tipo_persona, tasa_iva) AS (VALUES\n  ")
    out.write(",\n  ".join(values))
    out.write("\n),\n")

    out.write(
        "dilesa AS (SELECT id FROM core.empresas WHERE slug = 'dilesa'),\n"
        "personas_dilesa_match AS (\n"
        "  -- DISTINCT ON resuelve casos donde una persona tiene rows duplicadas\n"
        "  -- en erp.personas (legacy data). Prioriza el RFC sin guion (match exacto)\n"
        "  -- cuando existe; si no, toma cualquiera deterministicamente.\n"
        "  SELECT DISTINCT ON (e.rfc)\n"
        "    e.rfc AS excel_rfc, p.id AS persona_id\n"
        "  FROM excel e\n"
        "  JOIN erp.personas p ON replace(p.rfc, '-', '') = e.rfc\n"
        "  WHERE p.empresa_id = (SELECT id FROM dilesa)\n"
        "  ORDER BY e.rfc, (p.rfc = e.rfc) DESC, p.id\n"
        "),\n"
        "to_insert_personas AS (\n"
        "  SELECT e.codigo, e.nombre, e.rfc, e.tipo_persona\n"
        "  FROM excel e\n"
        "  WHERE NOT EXISTS (SELECT 1 FROM personas_dilesa_match m WHERE m.excel_rfc = e.rfc)\n"
        "),\n"
        "inserted_personas AS (\n"
        "  INSERT INTO erp.personas (empresa_id, nombre, rfc, tipo_persona)\n"
        "  SELECT (SELECT id FROM dilesa), tip.nombre, tip.rfc, tip.tipo_persona\n"
        "  FROM to_insert_personas tip\n"
        "  RETURNING id, rfc\n"
        "),\n"
        "all_personas AS (\n"
        "  SELECT excel_rfc AS rfc, persona_id FROM personas_dilesa_match\n"
        "  UNION ALL\n"
        "  SELECT rfc, id FROM inserted_personas\n"
        "),\n"
        "to_insert_proveedores AS (\n"
        "  SELECT e.codigo, ap.persona_id, e.tasa_iva\n"
        "  FROM excel e\n"
        "  JOIN all_personas ap ON ap.rfc = e.rfc\n"
        "  WHERE NOT EXISTS (\n"
        "    SELECT 1 FROM erp.proveedores pr\n"
        "    WHERE pr.persona_id = ap.persona_id\n"
        "      AND pr.empresa_id = (SELECT id FROM dilesa)\n"
        "      AND pr.deleted_at IS NULL\n"
        "  )\n"
        "),\n"
        "inserted_proveedores AS (\n"
        "  INSERT INTO erp.proveedores (empresa_id, persona_id, codigo, tasa_iva, activo)\n"
        "  SELECT (SELECT id FROM dilesa), tip.persona_id, tip.codigo, tip.tasa_iva, true\n"
        "  FROM to_insert_proveedores tip\n"
        "  RETURNING id\n"
        ")\n"
        "SELECT\n"
        "  (SELECT COUNT(*) FROM excel) AS rfcs_excel,\n"
        "  (SELECT COUNT(*) FROM personas_dilesa_match) AS ya_persona_dilesa,\n"
        "  (SELECT COUNT(*) FROM inserted_personas) AS personas_creadas,\n"
        "  (SELECT COUNT(*) FROM inserted_proveedores) AS proveedores_creados;\n\n"
    )

    out.write("COMMIT;\n")

    return 0


if __name__ == "__main__":
    sys.exit(main())
