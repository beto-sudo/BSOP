#!/usr/bin/env python3
"""
Traspaso Capa B — contratos de obra + estimaciones de las hojas de DETALLE de
los Excel de proyecto DILESA (LDLE, LDS) a `dilesa.contratos_construccion`
(tipo no-vivienda) + `dilesa.obra_estimaciones`. Iniciativa dilesa-contratos-obra.

DRY_RUN por default: detecta bloques de contrato por ANCLA (no por índice fijo),
extrae cabecera + tabla de estimaciones, RECONCILIA contra el "Total de
estimaciones" de cada hoja, clasifica las hojas/bloques que NO son contrato, y
escribe el JSON intermedio (/tmp/obra_contratos_traspaso.json) SIN tocar la DB.
La carga real se hace después, tras revisión de Beto, desde ese JSON.

  python3 scripts/import_dilesa_obra_contratos.py        # dry-run + reporte

Notas de modelado (ver docs/planning/dilesa-contratos-obra.md, ADR-038):
- Una hoja puede traer VARIOS contratos: tiled horizontal (bloques lado a lado)
  y/o apilado vertical (etapas una debajo de otra). Cada "Contrato"+"Anticipo/
  Por estimar" es un contrato con su propio anticipo/retención/estimaciones.
- anticipo_pct y retencion_pct se leen de las etiquetas ("Anticipo (50%)",
  "Total retención (5%)").
- Estimaciones: etiqueta de texto libre ("Anticipo","1","2A","3 y 4",
  "Finiquito","Amortizacion 60%",...); se preservan las NEGATIVAS (amortizaciones,
  notas de crédito) verbatim — el "Total de estimaciones" de la hoja las netea.
- IVA: los montos vienen c/IVA y el Excel casi nunca desglosa subtotal/IVA en
  detalle → monto_total = valor del Excel, subtotal/iva/iva_tasa null (decisión
  ADR-038: desglosar solo donde esté especificado, no inferir tasa).
- proveedor/título: BEST-EFFORT (se capturan las filas de contexto sobre el
  ancla + el nombre de la hoja). Beto revisa y normaliza antes de crear personas.
"""
import json
import os
import re
import sys
import unicodedata
from datetime import datetime, date

EMPRESA_DILESA = "f5942ed4-7a6b-4c39-af18-67b9fbf7f479"
ARCHIVOS = [
    {
        "path": os.path.expanduser("~/Downloads/Proyecto LDS.xlsx"),
        "proyecto_id": "a506b99f-1b6e-4024-a94a-59deaed48727",
        "proyecto_nombre": "Lomas del Sol",
        "prefijo": "LDS",
    },
    {
        "path": os.path.expanduser("~/Downloads/Proyecto LDLE.xlsx"),
        "proyecto_id": "42c64197-2358-4607-a21c-97556ceb3110",
        "proyecto_nombre": "Lomas de los Encinos",
        "prefijo": "LDLE",
    },
]

# Tolerancia de reconciliación: Σ estimaciones vs "Total de estimaciones" de la hoja.
RECON_TOL = 2.0  # pesos (los Excel tienen redondeos a centavos)


def norm(s):
    if s is None:
        return ""
    s = "".join(c for c in unicodedata.normalize("NFD", str(s)) if unicodedata.category(c) != "Mn")
    return s.strip().lower()


def num(v):
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    s = str(v).replace("$", "").replace(",", "").strip()
    if s in ("", "#DIV/0!", "#REF!", "-", "N/A", "n/a", "ok", "OK"):
        return None
    try:
        return round(float(s), 2)
    except ValueError:
        return None


def col_letter(j):
    """0-based índice -> letra(s) de columna estilo Excel (para source_ref)."""
    s = ""
    j += 1
    while j > 0:
        j, r = divmod(j - 1, 26)
        s = chr(65 + r) + s
    return s


def pct_from_label(label):
    """'Anticipo (50%)' -> 50.0 ; 'Total retención (5%)' -> 5.0 ; None si no hay %."""
    m = re.search(r"(\d+(?:\.\d+)?)\s*%", str(label))
    return float(m.group(1)) if m else None


_MESES = {
    "ene": 1, "feb": 2, "mar": 3, "mzo": 3, "abr": 4, "may": 5, "jun": 6,
    "jul": 7, "ago": 8, "sep": 9, "set": 9, "oct": 10, "nov": 11, "dic": 12,
}


def parse_fecha(v):
    """Devuelve (iso_or_None, raw_str). Acepta datetime, '16-Dic-21', '3dic25',
    '9ene26', '26-02-24', '20 ago 25', '21-mayo-25'. 'MAR'/None -> (None, raw)."""
    if v is None or v == "":
        return None, None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d"), v.isoformat()
    raw = str(v).strip()
    n = norm(raw).replace(" ", "")
    # dd-mmm-yy / ddmmmyy / dd-mmmmes-yy
    m = re.match(r"^(\d{1,2})[-/]?([a-z]{3,})[-/]?(\d{2,4})$", n)
    if m:
        d, mes, y = m.group(1), m.group(2)[:3], m.group(3)
        if mes in _MESES:
            y = int(y)
            y = y + 2000 if y < 100 else y
            try:
                return date(y, _MESES[mes], int(d)).strftime("%Y-%m-%d"), raw
            except ValueError:
                return None, raw
    # dd-mm-yy / dd-mm-yyyy
    m = re.match(r"^(\d{1,2})[-/](\d{1,2})[-/](\d{2,4})$", n)
    if m:
        d, mm, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        y = y + 2000 if y < 100 else y
        try:
            return date(y, mm, d).strftime("%Y-%m-%d"), raw
        except ValueError:
            return None, raw
    return None, raw  # no parseable -> se carga null, raw queda para revisión


# --- Palabras clave de cabecera de contrato ----------------------------------
ANCHOR_CONTRATO = {"contrato", "presupuesto"}
HEADER_LABELS = ("anticipo", "por estimar")


def is_anchor_row(rownorm):
    """¿Esta fila (lista de norms) contiene una o más anclas de contrato?
    Devuelve lista de índices de columna donde hay 'contrato'/'presupuesto' que
    tienen 'anticipo' o 'por estimar' a la derecha (dentro de 3 cols)."""
    cols = []
    for j, n in enumerate(rownorm):
        if n in ANCHOR_CONTRATO or (n.startswith("contrato") and len(n) < 14):
            # busca 'anticipo'/'por estimar' a la derecha
            right = " ".join(rownorm[j + 1 : j + 4])
            if any(h in right for h in HEADER_LABELS):
                cols.append(j)
    return cols


def find_anchor_group(rownorm, c):
    """Dado el col de 'Contrato' en c, ubica anticipo_col y porestimar_col."""
    anticipo_col = None
    porestimar_col = None
    anticipo_pct = None
    for j in range(c + 1, min(c + 4, len(rownorm))):
        n = rownorm[j]
        if "anticipo" in n and anticipo_col is None:
            anticipo_col = j
            anticipo_pct = pct_from_label(n)
        elif "por estimar" in n and porestimar_col is None:
            porestimar_col = j
    return anticipo_col, porestimar_col, anticipo_pct


def is_estim_subheaders(rownorm):
    """Una fila puede tener VARIAS tablas de estimaciones tiled lado a lado
    (ESTRELLA: 4 tablas Estimación|#Factura|Total en la misma fila). Devuelve la
    LISTA de (estim_col, total_col, factura_col, fecha_col), una por cada celda
    'estimacion'/'concepto', emparejada con el 'total' exacto más cercano a su
    derecha."""
    estim_cols = [j for j, n in enumerate(rownorm) if n in ("estimacion", "concepto")]
    total_cols = [j for j, n in enumerate(rownorm) if n == "total"]
    factura_cols = [j for j, n in enumerate(rownorm) if "factura" in n]
    fecha_cols = [j for j, n in enumerate(rownorm) if n in ("fecha", "mar")]
    out = []
    for ec in estim_cols:
        tc = next((t for t in total_cols if t > ec), None)
        if tc is None:
            continue
        fc = next((f for f in factura_cols if ec < f < tc), None)
        dc = next((d for d in reversed(fecha_cols) if d < ec), None)
        if dc is None:
            dc = ec - 1  # Fecha suele ir inmediatamente a la izquierda de Estimación
        out.append((ec, tc, fc, dc))
    return out


def classify_skipped(sheet, rows_norm):
    """Adivina la forma de una hoja sin ancla de contrato, para el reporte."""
    flat = " | ".join(" ".join(r) for r in rows_norm[:6])
    if "proveedor" in flat and "costo" in flat:
        return "lista de proveedores/compras (PROVEEDOR·COSTO·DESCRIPCION)"
    if "presupuesto previo" in flat or ("gasto real" in flat and "concepto" in flat):
        return "detalle tipo-RESUMEN (presupuesto vs gasto real) — Capa A"
    if "avance" in flat and ("aplicado" in flat or "ejecutado" in flat):
        return "avance de terracerías (volumen x PU, % avance) — no contrato"
    if "comparativo" in flat or "ib construcciones" in flat:
        return "comparativo de cotizaciones — no es contrato adjudicado"
    if "factibilidad" in flat or "licencia" in flat:
        return "trámites/licencias (montos sueltos)"
    return "sin ancla de contrato — revisar manualmente"


def parse_sheet(file_meta, sheet, rows):
    """Extrae contratos de una hoja. Devuelve (contratos, skipped_or_None)."""
    rows_norm = [[norm(c) for c in r] for r in rows]
    width = max((len(r) for r in rows_norm), default=0)
    rows_norm = [r + [""] * (width - len(r)) for r in rows_norm]

    # 1) anclas y sub-headers de estimaciones (con su fila)
    anchors = []  # (row_idx, contrato_col, anticipo_col, porestimar_col, anticipo_pct)
    for i, rn in enumerate(rows_norm):
        for c in is_anchor_row(rn):
            a_col, pe_col, a_pct = find_anchor_group(rn, c)
            anchors.append((i, c, a_col, pe_col, a_pct))
    subheaders = []  # (row_idx, estim_col, total_col, factura_col, fecha_col)
    for i, rn in enumerate(rows_norm):
        for sh in is_estim_subheaders(rn):
            subheaders.append((i, *sh))

    if not anchors:
        return [], {
            "file": file_meta["prefijo"],
            "sheet": sheet,
            "shape": classify_skipped(sheet, rows_norm),
            "filas": len(rows),
        }

    contratos = []
    anchors.sort()
    for idx, (arow, ccol, acol, pecol, apct) in enumerate(anchors):
        # límite inferior del bloque = siguiente ancla en una banda de columna que se solape
        next_row = len(rows_norm)
        for j, (arow2, ccol2, *_rest) in enumerate(anchors):
            if arow2 > arow and abs(ccol2 - ccol) <= 1:
                next_row = min(next_row, arow2)
                break

        def cell(r, c):
            return rows[r][c] if (0 <= r < len(rows) and c is not None and c < len(rows[r])) else None

        # valor del contrato y anticipo (fila inmediatamente debajo del ancla)
        valor_total = num(cell(arow + 1, ccol))
        anticipo_monto = num(cell(arow + 1, acol)) if acol is not None else None
        # algunos bloques traen "IVA INCLUIDO"/"con IVA" o folio en la col izquierda
        # de la fila de valores; el valor real puede estar en ccol de arow+1.

        # etiquetas de cabecera debajo del ancla (en ccol) -> retención/totales
        retencion_pct = None
        header_total_est = None
        header_total_pagado = None
        for r in range(arow + 1, min(arow + 9, next_row)):
            lbl = norm(cell(r, ccol))
            if "retencion" in lbl:
                retencion_pct = pct_from_label(cell(r, ccol))
            elif lbl == "total de estimaciones":
                header_total_est = num(cell(r, pecol))
            elif lbl == "total pagado":
                header_total_pagado = num(cell(r, pecol))

        # sub-header de estimaciones de este bloque: el más cercano debajo, en la
        # misma banda de columna (estim_col ~ contrato_col). Tiled horizontal: cada
        # bloque agarra la sub-tabla cuyo estim_col coincide con SU contrato_col.
        sh = None
        for (srow, ecol, tcol, fcol, dcol) in sorted(subheaders):
            if srow > arow and srow < next_row and abs(ecol - ccol) <= 1:
                sh = (srow, ecol, tcol, fcol, dcol)
                break

        estimaciones = []
        if sh:
            srow, ecol, tcol, fcol, dcol = sh
            # fin de la tabla: siguiente ancla o siguiente sub-header en banda, o blancos
            end = next_row
            for (srow2, ecol2, *_r) in sorted(subheaders):
                if srow2 > srow and abs(ecol2 - ccol) <= 1:
                    end = min(end, srow2)
                    break
            blanks = 0
            orden = 0
            for r in range(srow + 1, end):
                etiqueta_raw = cell(r, ecol)
                monto = num(cell(r, tcol))
                etqn = norm(etiqueta_raw)
                # corta si aparece otra etiqueta de cabecera (otro bloque pegado)
                if etqn in ("contrato", "presupuesto") or etqn.startswith("total retencion"):
                    break
                if (etiqueta_raw is None or etqn == "") and monto is None:
                    blanks += 1
                    if blanks >= 3:
                        break
                    continue
                blanks = 0
                fiso, fraw = parse_fecha(cell(r, dcol))
                # nota de pago: texto en la col a la derecha de Total ("pagada","pag 13oct")
                nota = cell(r, tcol + 1)
                nota = str(nota).strip() if (nota not in (None, "") and num(nota) is None) else None
                orden += 1
                estimaciones.append({
                    "orden": orden,
                    "etiqueta": (str(etiqueta_raw).strip() if etiqueta_raw is not None else None),
                    "fecha_iso": fiso,
                    "fecha_raw": fraw,
                    "factura_ref": clean_ref(cell(r, fcol)),
                    "monto_total": monto,
                    "es_anticipo": etqn.startswith("anticipo"),
                    "es_finiquito": etqn.startswith("finiquito") or etqn.startswith("finquito") or "liquidacion" in etqn,
                    "nota_pago": nota,
                    "source_ref": f"{os.path.basename(file_meta['path'])}/{sheet}/{col_letter(ecol)}{r+1}",
                })

        # contexto: filas de texto sobre el ancla en ccol (proveedor/título/etapa)
        context = []
        for r in range(max(0, arow - 4), arow):
            for c in (ccol, ccol - 1, 0):
                v = cell(r, c)
                if v is not None and str(v).strip() and norm(v) not in ANCHOR_CONTRATO:
                    context.append(str(v).strip())
                    break
        # proveedor_guess = primera línea de contexto que parezca nombre (heurística)
        proveedor_guess = context[0] if context else None
        titulo = " / ".join(context[1:]) if len(context) > 1 else (context[0] if context else sheet)

        # Σ incluye el anticipo y las amortizaciones (negativas) → reconcilia contra
        # "Total Pagado" (efectivo desembolsado), no contra "Total de estimaciones"
        # (que excluye el anticipo). Fallback a Total de estimaciones si no hay pagado.
        suma = round(sum(e["monto_total"] or 0 for e in estimaciones), 2)
        recon_target = header_total_pagado if header_total_pagado is not None else header_total_est
        recon_basis = "pagado" if header_total_pagado is not None else ("estimaciones" if header_total_est is not None else None)
        recon_diff = round(suma - recon_target, 2) if recon_target is not None else None
        recon_ok = (recon_diff is not None and abs(recon_diff) <= RECON_TOL)

        contratos.append({
            "empresa_id": EMPRESA_DILESA,
            "proyecto_id": file_meta["proyecto_id"],
            "proyecto_nombre": file_meta["proyecto_nombre"],
            "file": file_meta["prefijo"],
            "sheet": sheet,
            "anchor_cell": f"{col_letter(ccol)}{arow+1}",
            "proveedor_guess": proveedor_guess,
            "titulo": titulo,
            "context_rows": context,
            "anticipo_pct": apct,
            "retencion_pct": retencion_pct,
            "valor_total": valor_total,
            "anticipo_monto": anticipo_monto,
            "header_total_estimaciones": header_total_est,
            "header_total_pagado": header_total_pagado,
            "n_estimaciones": len(estimaciones),
            "suma_estimaciones": suma,
            "recon_target": recon_target,
            "recon_basis": recon_basis,
            "recon_diff": recon_diff,
            "recon_ok": recon_ok,
            "estimaciones": estimaciones,
        })

    return contratos, None


# =============================================================================
# CARGA A PROD (modo --emit-sql) — decisiones cerradas con Beto 2026-06-01:
#  - Mapear contratistas existentes; crear los 5 con nombre; los sin contratista
#    a UN placeholder reasignable ("CONTRATISTA POR ASIGNAR — OBRA"), con el hint
#    del Excel en notas para reasignar en la UI.
#  - SIMAS se modela como contrato (convenio urbanización, 15 pagos programados).
# =============================================================================
PLACEHOLDER = "CONTRATISTA POR ASIGNAR — OBRA"

# Contratistas a CREAR en erp.personas (no existían). (nombre, tipo_persona)
NEW_CONTRATISTAS = [
    ("ROMEO GONZALEZ", "fisica"),
    ("WILLYSONS CONSTRUCCIONES", "moral"),
    ("MIGUEL ANGEL QUINTERO FUENTES", "fisica"),
    ("MAYRA MARGARITA PEREZ ARENAS", "fisica"),
    ("ESTRELLA", "fisica"),  # razón social a confirmar; placeholder de tipo
    (PLACEHOLDER, "fisica"),
]

# Resolución contrato → contratista, por (prefijo, hoja, celda-ancla). Determinístico.
# Los que ya existen en erp.personas se matchean por nombre exacto (subquery).
SIMAS_CONTRATISTA = "SISTEMA MUNICIPAL DE AGUAS Y SANEAMIENTO DE PIEDRAS NEGRAS, COAHUILA"
RESOLUTION = {
    ("LDS", "ESTRELLA", "B3"): "ESTRELLA",
    ("LDS", "ESTRELLA", "G3"): "ESTRELLA",
    ("LDS", "ESTRELLA", "L3"): "ESTRELLA",
    ("LDS", "ESTRELLA", "P3"): "ESTRELLA",
    ("LDS", "BARDA", "B3"): PLACEHOLDER,
    ("LDS", "BARDA2", "B3"): PLACEHOLDER,
    ("LDS", "BANQUETA", "B3"): PLACEHOLDER,
    ("LDS", "BANQUETA", "K3"): PLACEHOLDER,
    ("LDS", "CASETA", "B3"): PLACEHOLDER,
    ("LDS", "PAVIMENTACION", "C4"): PLACEHOLDER,
    ("LDS", "PORTON", "C4"): "EMMA LIZETHE CAZARES SANTILLAN",
    ("LDS", "PORTON", "I4"): "TELECOMUNICACIONES DE COAHUILA",
    ("LDS", "ELECTRIFICACION", "C4"): PLACEHOLDER,
    ("LDS", "ELECTRIFICACION", "H4"): PLACEHOLDER,
    ("LDLE", "ELECTRIFICACION", "C4"): "ROMEO GONZALEZ",
    ("LDLE", "ELECTRIFICACION", "I4"): "ELECTROGAZA",
    ("LDLE", "ELECTRIFICACION", "C26"): "ELECTROGAZA",
    ("LDLE", "ELECTRIFICACION", "C65"): "ELECTROGAZA",
    ("LDLE", "ELECTRIFICACION", "I65"): "ELECTROGAZA",
    ("LDLE", "ELECTRIFICACION", "B81"): "ELECTROGAZA",
    ("LDLE", "AGUA POTABLE DRENAJE", "B3"): "WILLYSONS CONSTRUCCIONES",
    ("LDLE", "AGUA POTABLE DRENAJE", "G3"): "WILLYSONS CONSTRUCCIONES",
    ("LDLE", "PAVIMENTACION", "C4"): "MATERIALES SAN RODRIGO",
    ("LDLE", "PAVIMENTACION", "I4"): "MATERIALES SAN RODRIGO",
    ("LDLE", "PAVIMENTACION", "O4"): "MATERIALES SAN RODRIGO",
    ("LDLE", "PAVIMENTACION", "S4"): "MATERIALES SAN RODRIGO",
    ("LDLE", "MONOLITO", "C4"): "MIGUEL ANGEL QUINTERO FUENTES",
    ("LDLE", "NOMENCLATURA", "C5"): "MAYRA MARGARITA PEREZ ARENAS",
    ("LDLE", "CORDON", "C4"): "MIGUEL ANGEL QUINTERO FUENTES",
    ("LDLE", "URBANIZACIÓN", "C5"): "TUBOS Y CONEXIONES DE COAHUILA",
    ("LDLE", "VANDALIZADAS", "C4"): "EMANUEL FIDENCIO MORADO DE LUNA",
}

TIPO_OBRA_CABECERA = {"BARDA", "BARDA2", "BANQUETA", "CASETA", "PORTON", "MONOLITO", "NOMENCLATURA"}
TIPO_TAREA_MENOR = {"VANDALIZADAS"}


def tipo_obra(sheet):
    if sheet in TIPO_OBRA_CABECERA:
        return "obra_cabecera"
    if sheet in TIPO_TAREA_MENOR:
        return "tarea_menor"
    return "urbanizacion"  # ELECTRIFICACION, AGUA POTABLE, PAVIMENTACION, CORDON, URBANIZACIÓN, ESTRELLA, SIMAS


def resolve_contratista(ct):
    return RESOLUTION.get((ct["file"], ct["sheet"], ct["anchor_cell"]), PLACEHOLDER)


def parse_simas(fm, rows):
    """SIMAS = convenio de derechos de interconexión de agua. No tiene ancla de
    contrato; se modela a mano: valor=Convenio, 15 pagos programados como
    estimaciones (los que traen monto en el Excel = pagados; el resto = programado)."""
    valor = None
    for row in rows:
        for j, v in enumerate(row):
            if norm(v) == "convenio":
                valor = next((num(x) for x in row[j + 1:] if num(x) is not None), None)
    pagos = []
    for row in rows:
        a = num(row[0]) if row else None
        if a is not None and 1 <= a <= 40 and float(a).is_integer():
            fiso, fraw = parse_fecha(row[1] if len(row) > 1 else None)
            monto_d = num(row[3]) if len(row) > 3 else None
            pagos.append((int(a), fiso, fraw, monto_d))
    if not pagos:
        return None
    cuota = round(valor / len(pagos), 2) if valor else (pagos[0][3] or 0)
    estimaciones = []
    for (n, fiso, fraw, monto_d) in pagos:
        estimaciones.append({
            "orden": n,
            "etiqueta": f"Pago {n}",
            "fecha_iso": fiso,
            "fecha_raw": fraw,
            "factura_ref": None,
            "monto_total": monto_d if monto_d is not None else cuota,
            "es_anticipo": False,
            "es_finiquito": (n == len(pagos)),
            "nota_pago": "pagada" if monto_d is not None else "programado",
            "source_ref": f"{os.path.basename(fm['path'])}/SIMAS/r{n}",
        })
    suma = round(sum(e["monto_total"] or 0 for e in estimaciones), 2)
    return {
        "empresa_id": EMPRESA_DILESA, "proyecto_id": fm["proyecto_id"],
        "proyecto_nombre": fm["proyecto_nombre"], "file": fm["prefijo"], "sheet": "SIMAS",
        "anchor_cell": "B6", "proveedor_guess": SIMAS_CONTRATISTA, "titulo": "Derechos de interconexión de agua (convenio)",
        "context_rows": ["SIMAS", "Derechos de interconexión"], "anticipo_pct": None, "retencion_pct": None,
        "valor_total": valor, "anticipo_monto": None, "header_total_estimaciones": None,
        "header_total_pagado": None, "n_estimaciones": len(estimaciones), "suma_estimaciones": suma,
        "recon_target": valor, "recon_basis": "convenio",
        "recon_diff": round(suma - valor, 2) if valor else None,
        "recon_ok": (valor is not None and abs(suma - valor) <= RECON_TOL), "estimaciones": estimaciones,
    }


def _s(v):
    """Literal SQL: texto escapado, NULL, número o bool."""
    if v is None:
        return "NULL"
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        return repr(round(v, 2))
    return "'" + str(v).replace("'", "''") + "'"


def clean_ref(v):
    """Referencia de factura/OC: int-float -> '4578' (no '4578.0'); resto -> texto."""
    if v is None or v == "":
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def build_load_sql(contratos):
    """SQL idempotente y transaccional para cargar contratos + estimaciones."""
    E = _s(EMPRESA_DILESA)
    out = ["BEGIN;", ""]

    # 1) Contratistas faltantes (idempotente por nombre exacto, case-insensitive).
    out.append("-- 1) Contratistas nuevos en erp.personas (idempotente por nombre)")
    vals = ",\n  ".join(
        f"({_s(n)}, {_s(tp)})" for (n, tp) in NEW_CONTRATISTAS
    )
    out.append(
        "INSERT INTO erp.personas (empresa_id, nombre, tipo, tipo_persona, activo)\n"
        f"SELECT {E}, v.nombre, 'contratista', v.tp, true\n"
        f"FROM (VALUES\n  {vals}\n) AS v(nombre, tp)\n"
        "WHERE NOT EXISTS (SELECT 1 FROM erp.personas p\n"
        "  WHERE p.deleted_at IS NULL AND upper(p.nombre) = upper(v.nombre));")
    out.append("")

    # 2) Limpiar carga previa de Capa B (idempotente).
    out.append("-- 2) Limpieza idempotente de carga previa de Capa B")
    out.append("DELETE FROM dilesa.obra_estimaciones WHERE source_ref LIKE 'Proyecto LD%';")
    out.append(f"DELETE FROM dilesa.contratos_construccion WHERE codigo LIKE 'OBRA-%' AND empresa_id = {E};")
    out.append("")

    # 3) Contratos (contratista resuelto por LATERAL al nombre).
    out.append("-- 3) Contratos de obra (tipo no-vivienda)")
    crows = []
    file_min = {}
    for ct in contratos:
        ds = [e["fecha_iso"] for e in ct["estimaciones"] if e["fecha_iso"]]
        if ds:
            file_min[ct["file"]] = min(file_min.get(ct["file"], "9999"), min(ds))
    for ct in contratos:
        codigo = f"OBRA-{ct['file']}-{ct['sheet'].replace(' ', '_')}-{ct['anchor_cell']}"
        ds = [e["fecha_iso"] for e in ct["estimaciones"] if e["fecha_iso"]]
        fecha = min(ds) if ds else file_min.get(ct["file"], "2022-01-01")
        contratista = resolve_contratista(ct) if ct["sheet"] != "SIMAS" else SIMAS_CONTRATISTA
        # anticipo_pct: el indicado, o calculado del monto si existe
        apct = ct["anticipo_pct"]
        if apct is None and ct.get("anticipo_monto") and ct.get("valor_total"):
            apct = round(ct["anticipo_monto"] / ct["valor_total"] * 100, 2)
        nota = f"Traspaso Capa B {ct['file']}/{ct['sheet']}."
        if contratista == PLACEHOLDER:
            nota += f" Contratista por asignar (Excel: '{ct['proveedor_guess']}' — {ct['titulo']})."
        if not ct["recon_ok"]:
            nota += f" ⚠️ Σ estimaciones ({ct['suma_estimaciones']}) ≠ total Excel ({ct['recon_target']})."
        crows.append(
            f"({_s(codigo)}, {_s(fecha)}, {_s(ct['proyecto_id'])}, {_s(ct['valor_total'] or 0)}, "
            f"{_s(tipo_obra(ct['sheet']))}, {_s(apct or 0)}, {_s(ct['retencion_pct'] or 0)}, "
            f"{_s(contratista)}, {_s(nota)})")
    crows_sql = ",\n  ".join(crows)
    out.append(
        "INSERT INTO dilesa.contratos_construccion\n"
        "  (empresa_id, codigo, fecha_contrato, contratista_id, proyecto_id, valor_total, tipo, anticipo_pct, retencion_pct, notas)\n"
        f"SELECT {E}, v.codigo, v.fecha::date, p.id, v.proyecto_id::uuid, v.valor, v.tipo, v.ant, v.ret, v.notas\n"
        f"FROM (VALUES\n  {crows_sql}\n) AS v(codigo, fecha, proyecto_id, valor, tipo, ant, ret, contratista, notas)\n"
        "JOIN LATERAL (SELECT id FROM erp.personas WHERE deleted_at IS NULL\n"
        "  AND upper(nombre) = upper(v.contratista) ORDER BY created_at LIMIT 1) p ON true;")
    out.append("")

    # 4) Estimaciones (contrato resuelto por codigo).
    out.append("-- 4) Estimaciones de monto")
    erows = []
    for ct in contratos:
        codigo = f"OBRA-{ct['file']}-{ct['sheet'].replace(' ', '_')}-{ct['anchor_cell']}"
        for e in ct["estimaciones"]:
            etiqueta = e["etiqueta"] or f"Estimación {e['orden']}"
            erows.append(
                f"({_s(codigo)}, {_s(etiqueta)}, {_s(e['orden'])}, {_s(e['fecha_iso'])}, "
                f"{_s(e['factura_ref'])}, {_s(e['monto_total'] or 0)}, {_s(e['es_anticipo'])}, "
                f"{_s(e['es_finiquito'])}, {_s(e['nota_pago'])}, {_s(e['source_ref'])})")
    erows_sql = ",\n  ".join(erows)
    out.append(
        "INSERT INTO dilesa.obra_estimaciones\n"
        "  (empresa_id, contrato_id, etiqueta, orden, fecha, factura_ref, monto_total, es_anticipo, es_finiquito, nota_pago, source_ref)\n"
        f"SELECT {E}, c.id, e.etiqueta, e.orden, e.fecha::date, e.factura_ref, e.monto, e.es_ant, e.es_fin, e.nota, e.source_ref\n"
        f"FROM (VALUES\n  {erows_sql}\n) AS e(codigo, etiqueta, orden, fecha, factura_ref, monto, es_ant, es_fin, nota, source_ref)\n"
        f"JOIN dilesa.contratos_construccion c ON c.codigo = e.codigo AND c.empresa_id = {E} AND c.deleted_at IS NULL;")
    out.append("")
    out.append("NOTIFY pgrst, 'reload schema';")
    out.append("COMMIT;")
    return "\n".join(out)


def main():
    import openpyxl

    all_contratos = []
    skipped = []
    print("=" * 78)
    for fm in ARCHIVOS:
        wb = openpyxl.load_workbook(fm["path"], data_only=True)
        print(f"\n■■■ {fm['prefijo']} → '{fm['proyecto_nombre']}'  ({fm['path'].split('/')[-1]})")
        for sheet in wb.sheetnames:
            if sheet == "RESUMEN":
                continue
            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            if sheet == "SIMAS":  # convenio de derechos — handler dedicado (decisión Beto)
                ct = parse_simas(fm, rows)
                if ct:
                    print(f"  {'OK ' if ct['recon_ok'] else '⚠️ '}{sheet:<22} {ct['anchor_cell']:<5} "
                          f"valor={_money(ct['valor_total']):>16} est={ct['n_estimaciones']:>2} "
                          f"Σ={_money(ct['suma_estimaciones']):>16} (convenio · {SIMAS_CONTRATISTA[:24]}…)")
                    all_contratos.append(ct)
                continue
            contratos, skip = parse_sheet(fm, sheet, rows)
            if skip:
                skipped.append(skip)
                continue
            for ct in contratos:
                flag = "OK " if ct["recon_ok"] else ("?? " if ct["recon_target"] is None else "⚠️ ")
                print(f"  {flag}{sheet:<22} {ct['anchor_cell']:<5} "
                      f"ant={str(ct['anticipo_pct'] or '-'):>4}% ret={str(ct['retencion_pct'] or '-'):>4}% "
                      f"valor={_money(ct['valor_total']):>16} "
                      f"est={ct['n_estimaciones']:>2} Σ={_money(ct['suma_estimaciones']):>16} "
                      f"recon={_money(ct['recon_diff'])}")
                if ct["proveedor_guess"]:
                    print(f"       proveedor?: {ct['proveedor_guess']!r}  | titulo: {ct['titulo']!r}")
            all_contratos.extend(contratos)
        wb.close()

    # --- reporte de hojas saltadas ---
    print(f"\n{'=' * 78}\nHOJAS NO-CONTRATO (saltadas, para decidir):")
    for s in skipped:
        print(f"  - {s['file']:<5} {s['sheet']:<22} ({s['filas']} filas) → {s['shape']}")

    # --- contratistas distintos (para resolución a erp.personas) ---
    provs = sorted({c["proveedor_guess"] for c in all_contratos if c["proveedor_guess"]})
    print(f"\nCONTRATISTAS/PROVEEDORES distintos detectados ({len(provs)}):")
    for p in provs:
        print(f"  - {p}")

    # --- reconciliación global ---
    n_ok = sum(1 for c in all_contratos if c["recon_ok"])
    n_warn = sum(1 for c in all_contratos if not c["recon_ok"] and c["recon_target"] is not None)
    n_noref = sum(1 for c in all_contratos if c["recon_target"] is None)
    n_est = sum(c["n_estimaciones"] for c in all_contratos)
    print(f"\n{'=' * 78}")
    print(f"CONTRATOS: {len(all_contratos)}  |  estimaciones: {n_est}")
    print(f"  reconcilian (±${RECON_TOL}): {n_ok}   ⚠️ difieren: {n_warn}   ?? sin total de ref: {n_noref}")
    if n_warn:
        print("\n  ⚠️ Contratos que NO reconcilian (revisar parser/Excel):")
        for c in all_contratos:
            if not c["recon_ok"] and c["recon_target"] is not None:
                print(f"     {c['file']} {c['sheet']} {c['anchor_cell']}: "
                      f"Σ={_money(c['suma_estimaciones'])} vs total={_money(c['recon_target'])} "
                      f"(dif {_money(c['recon_diff'])})")

    out_path = "/tmp/obra_contratos_traspaso.json"
    with open(out_path, "w") as f:
        json.dump({"contratos": all_contratos, "skipped": skipped}, f, ensure_ascii=False, indent=2)
    print(f"\nJSON intermedio → {out_path}  (DRY-RUN: nada escrito en la DB)")

    if "--emit-sql" in sys.argv:
        sql = build_load_sql(all_contratos)
        sql_path = "/tmp/obra_contratos_load.sql"
        with open(sql_path, "w") as f:
            f.write(sql)
        print(f"SQL de carga → {sql_path}  ({len(sql):,} bytes, {sql.count(chr(10))} líneas)")


def _money(v):
    return "—" if v is None else f"${v:,.2f}"


if __name__ == "__main__":
    main()
